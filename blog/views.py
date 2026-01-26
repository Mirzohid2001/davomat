from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from django.db.models import Q, Count, F
from django.utils import timezone
from datetime import timedelta, date
from .models import Employee, Attendance, DayOff, AttendanceImportLog, MonthlyEmployeeStat
from .forms import EmployeeForm, AttendanceForm, BulkAttendanceForm, AttendanceImportForm, DayOffForm
from django.forms import modelformset_factory
from django.db import transaction

import pandas as pd
import openpyxl
from django.http import HttpResponse
from .services import calculate_monthly_stats, calculate_working_days_in_month
from openpyxl.utils import get_column_letter

from django import forms
from decimal import Decimal
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule

class SalaryStatFilterForm(forms.Form):
    year = forms.IntegerField(label="Yil", min_value=2000, max_value=2100)
    month = forms.IntegerField(label="Oy", min_value=1, max_value=12)

class SalaryStatEditForm(forms.ModelForm):
    class Meta:
        model = MonthlyEmployeeStat
        fields = ['salary', 'currency', 'paid', 'bonus']
        widgets = {
            'salary': forms.NumberInput(attrs={'class': 'form-control', 'step': 'any'}),
            'currency': forms.Select(attrs={'class': 'form-select'}),
            'paid': forms.NumberInput(attrs={'class': 'form-control', 'step': 'any'}),
            'bonus': forms.NumberInput(attrs={'class': 'form-control', 'step': 'any'}),
        }

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, "Muvaffaqiyatli tizimga kirdingiz!")
            return redirect('dashboard')
        else:
            messages.error(request, "Login yoki parol noto'g'ri!")
    else:
        form = AuthenticationForm()
    return render(request, 'attendance/login.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.success(request, "Tizimdan chiqdingiz.")
    return redirect('login')

@login_required
def dashboard(request):
    today = date.today()
    employees = Employee.objects.filter(is_active=True)
    attendance_today = Attendance.objects.filter(date=today)
    
    # Бугун ёпиқ кун ёки якшанбами текшириш
    is_dayoff = DayOff.objects.filter(date=today).exists()
    is_sunday = today.weekday() == 6
    
    # Офис ходимларини ва ёпиқ кунларда барчани чиқариш
    if is_dayoff or is_sunday:
        # Ёпиқ кунларда ҳеч кимни "киритилмаган"га чиқармаслик
        not_filled = Employee.objects.none()
    else:
        # Фақат офис ходимларидан бошқаларни текшириш (office емас ходимлар)
        employees_need_attendance = employees.exclude(employee_type='office')
        not_filled = employees_need_attendance.exclude(id__in=attendance_today.values_list('employee', flat=True))
    
    stats = Attendance.objects.values('status').annotate(count=Count('id'))
    best_employees = Attendance.objects.filter(
        status='present',
        date__gte=today - timedelta(days=30)
    ).values('employee__last_name', 'employee__first_name').annotate(
        present_count=Count('id')
    ).order_by('-present_count')[:5]
    return render(request, 'attendance/dashboard.html', {
        'today': today,
        'employees': employees,
        'attendance_today': attendance_today,
        'not_filled': not_filled,
        'stats': stats,
        'best_employees': best_employees,
    })

@login_required
def employee_list(request):
    search = request.GET.get('q', '')
    employees = Employee.objects.all()
    if search:
        employees = employees.filter(
            Q(first_name__icontains=search) | 
            Q(last_name__icontains=search) | 
            Q(position__icontains=search) | 
            Q(department__icontains=search) | 
            Q(phone_number__icontains=search)
        )
    return render(request, 'attendance/employees.html', {'employees': employees, 'search': search})

@login_required
def employee_create(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Xodim muvaffaqiyatli qo'shildi!")
            return redirect('employee_list')
    else:
        form = EmployeeForm()
    return render(request, 'attendance/employee_form.html', {'form': form})

@login_required
def employee_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, "Xodim ma'lumoti yangilandi!")
            return redirect('employee_list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'attendance/employee_form.html', {'form': form, 'employee': employee})

@login_required
def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        messages.success(request, "Xodim o'chirildi!")
        return redirect('employee_list')
    return render(request, 'attendance/employee_confirm_delete.html', {'employee': employee})


@login_required
def bulk_attendance_create(request):
    today = date.today()
    # Бугун ёпиқ кун ёки якшанбами текшириш
    is_dayoff = DayOff.objects.filter(date=today).exists()
    is_sunday = today.weekday() == 6
    
    if is_dayoff or is_sunday:
        dayoff_reason = DayOff.objects.filter(date=today).first()
        reason = dayoff_reason.reason if dayoff_reason else "Якшанба"
        messages.error(request, f"Бугун {reason} - давомат киритиб бўлмайди!")
        return redirect('dashboard')
    
    employees = Employee.objects.filter(is_active=True)
    date_val = request.GET.get('date') or date.today()
    AttendanceFormSet = modelformset_factory(
        Attendance, form=AttendanceForm, can_delete=False,
        fields=['employee', 'date', 'status', 'comment', 'attachment'],
        extra=employees.count()
    )

    attendance_today = Attendance.objects.filter(date=date_val)
    initial_data = []
    employee_id_to_fio = {}
    for emp in employees:
        employee_id_to_fio[str(emp.id)] = {
            "first_name": emp.first_name,
            "last_name": emp.last_name,
        }
        rec = attendance_today.filter(employee=emp).first()
        if rec:
            initial_data.append({
                'employee': rec.employee.id,
                'date': rec.date,
                'status': rec.status,
                'comment': rec.comment,
                'attachment': rec.attachment
            })
        else:
            initial_data.append({
                'employee': emp.id,
                'date': date_val,
                'status': 'present',
            })

    if request.method == 'POST':
        formset = AttendanceFormSet(request.POST, request.FILES,
                                    queryset=Attendance.objects.none(),
                                    initial=initial_data)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Davomat muvaffaqiyatli saqlandi!")
            return redirect('attendance_list')
    else:
        formset = AttendanceFormSet(queryset=Attendance.objects.none(), initial=initial_data)

    dayoff = DayOff.objects.filter(date=date_val).first() if hasattr(DayOff, 'date') else None

    return render(request, 'attendance/bulk_attendance_form.html', {
        'formset': formset,
        'employee_id_to_fio': employee_id_to_fio,
        'date_val': date_val,
        'dayoff': dayoff,
    })



@login_required
def attendance_list(request):
    day = request.GET.get('date')
    employees = Employee.objects.filter(is_active=True)
    filters = {}
    if day:
        try:
            day = date.fromisoformat(day)
            filters['date'] = day
        except Exception:
            day = date.today()
            filters['date'] = day
    else:
        day = date.today()
        filters['date'] = day
    if request.GET.get('status'):
        filters['status'] = request.GET['status']
    if request.GET.get('department'):
        filters['employee__department'] = request.GET['department']
    if request.GET.get('position'):
        filters['employee__position'] = request.GET['position']
    attendance = Attendance.objects.filter(**filters).select_related('employee')
    return render(request, 'attendance/attendance_list.html', {
        'attendance': attendance,
        'today': day,
        'employees': employees,
    })

@login_required
def attendance_update(request, pk):
    record = get_object_or_404(Attendance, pk=pk)
    if request.method == 'POST':
        form = AttendanceForm(request.POST, request.FILES, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, "Davomat yangilandi!")
            return redirect('attendance_list')
    else:
        form = AttendanceForm(instance=record)
    return render(request, 'attendance/attendance_form.html', {'form': form, 'record': record})

@login_required
def attendance_delete(request, pk):
    record = get_object_or_404(Attendance, pk=pk)
    if request.method == 'POST':
        record.delete()
        messages.success(request, "Davomat o'chirildi!")
        return redirect('attendance_list')
    return render(request, 'attendance/attendance_confirm_delete.html', {'record': record})

@login_required
def attendance_import(request):
    errors = []
    if request.method == 'POST':
        form = AttendanceImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            ext = file.name.split('.')[-1]
            try:
                if ext in ['xls', 'xlsx']:
                    df = pd.read_excel(file)
                else:
                    df = pd.read_csv(file)
                count = 0
                for idx, row in df.iterrows():
                    try:
                        emp = Employee.objects.filter(
                            last_name=row['last_name'],
                            first_name=row['first_name']
                        ).first()
                        if not emp:
                            errors.append(f"{idx+2}-satr: Xodim topilmadi ({row.get('last_name','')} {row.get('first_name','')})")
                            continue
                        # Sana formati tekshiruvi
                        try:
                            row_date = row['date']
                            if not isinstance(row_date, str):
                                row_date = str(row_date)
                            date_obj = date.fromisoformat(row_date)
                        except Exception:
                            errors.append(f"{idx+2}-satr: Sana formati noto'g'ri ({row.get('date','')})")
                            continue
                        Attendance.objects.update_or_create(
                            employee=emp,
                            date=date_obj,
                            defaults={
                                'status': row['status'],
                                'comment': row.get('comment', ''),
                            }
                        )
                        count += 1
                    except Exception as e:
                        errors.append(f"{idx+2}-satr: {str(e)}")
                AttendanceImportLog.objects.create(
                    user=request.user,
                    file_name=file.name,
                    record_count=count,
                    success=(len(errors) == 0),
                    log='; '.join(errors) if errors else 'OK'
                )
                if count:
                    messages.success(request, f"{count} ta davomat import qilindi!")
                if errors:
                    messages.error(request, f"Quyidagi satrlarda xatoliklar: {'; '.join(errors)}")
            except Exception as e:
                AttendanceImportLog.objects.create(
                    user=request.user,
                    file_name=file.name,
                    record_count=0,
                    success=False,
                    log=str(e)
                )
                messages.error(request, f"Importda xatolik: {str(e)}")
            return redirect('attendance_import')
    else:
        form = AttendanceImportForm()
    return render(request, 'attendance/attendance_import.html', {'form': form})

@login_required
def attendance_export(request):
    if request.method == 'POST':
        # Get filter parameters
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
        department = request.POST.get('department')
        status = request.POST.get('status')
        
        # Build query
        filters = {}
        if date_from and date_to:
            try:
                date_from_obj = date.fromisoformat(date_from)
                date_to_obj = date.fromisoformat(date_to)
                filters['date__range'] = (date_from_obj, date_to_obj)
            except Exception as e:
                messages.error(request, f"Sana formatida xatolik: {str(e)}")
                return redirect('attendance_export')
        if department:
            filters['employee__department'] = department
        if status:
            filters['status'] = status
            
        # Get data
        attendances = Attendance.objects.filter(**filters).select_related('employee')
        
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Davomat"
        # Add header row
        headers = ['Sana', 'Familiya', 'Ismi', 'Lavozim', 'Bo\'lim', 'Status', 'Izoh']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Add data rows
        for row_num, att in enumerate(attendances, 2):
            ws.cell(row=row_num, column=1, value=att.date)
            ws.cell(row=row_num, column=2, value=att.employee.last_name)
            ws.cell(row=row_num, column=3, value=att.employee.first_name)
            ws.cell(row=row_num, column=4, value=att.employee.position)
            ws.cell(row=row_num, column=5, value=att.employee.department)
            ws.cell(row=row_num, column=6, value=att.get_status_display())
            ws.cell(row=row_num, column=7, value=att.comment)
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="attendance.xlsx"'
        wb.save(response)
        return response
    else:
        # Get unique departments for filter options
        departments = Employee.objects.values_list('department', flat=True).distinct()
        # Get today and start of month for default date range
        today = timezone.now().date()
        start_of_month = today.replace(day=1)
        
        context = {
            'departments': departments,
            'statuses': Attendance.STATUS_CHOICES,
            'today': today,
            'start_of_month': start_of_month,
        }
        return render(request, 'attendance/attendance_export.html', context)

@login_required
def dayoff_list(request):
    days = DayOff.objects.order_by('-date')
    return render(request, 'attendance/dayoff_list.html', {'days': days})

@login_required
def dayoff_create(request):
    if request.method == 'POST':
        form = DayOffForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Yopiq sana qo'shildi!")
            return redirect('dayoff_list')
    else:
        form = DayOffForm()
    return render(request, 'attendance/dayoff_form.html', {'form': form})

@login_required
def dayoff_delete(request, pk):
    day = get_object_or_404(DayOff, pk=pk)
    if request.method == 'POST':
        day.delete()
        messages.success(request, "Yopiq sana o'chirildi!")
        return redirect('dayoff_list')
    return render(request, 'attendance/dayoff_confirm_delete.html', {'day': day})

@login_required
def individual_attendance_create(request, employee_id=None):
    """Har bir xodim uchun alohida davomat kiritish"""
    if employee_id:
        employee = get_object_or_404(Employee, pk=employee_id)
    else:
        # Agar employee_id ko'rsatilmagan bo'lsa, xodimni tanlash uchun forma ko'rsatiladi
        if request.method == 'POST':
            employee_id = request.POST.get('employee')
            if employee_id:
                return redirect('individual_attendance_create', employee_id=employee_id)
            else:
                messages.error(request, "Xodimni tanlang!")
        
        employees = Employee.objects.filter(is_active=True).order_by('location', 'department', 'last_name')
        locations = Employee.LOCATION_CHOICES
        return render(request, 'attendance/select_employee.html', {
            'employees': employees,
            'locations': locations
        })
    
    # Tanlangan xodim uchun davomat kiritish
    try:
        date_val = request.GET.get('date')
        if date_val:
            date_val = date.fromisoformat(date_val)
        else:
            date_val = date.today()
    except ValueError:
        date_val = date.today()
        messages.warning(request, "Noto'g'ri sana formati. Bugungi sana ishlatilmoqda.")
    
    # Танланган санани текшириш
    is_dayoff = DayOff.objects.filter(date=date_val).exists()
    is_sunday = date_val.weekday() == 6
    
    if is_dayoff or is_sunday:
        dayoff_reason = DayOff.objects.filter(date=date_val).first()
        reason = dayoff_reason.reason if dayoff_reason else "Якшанба"
        messages.error(request, f"{date_val.strftime('%d.%m.%Y')} - {reason} кунида давомат киритиб бўлмайди!")
        return redirect('dashboard')
    
    attendance = Attendance.objects.filter(employee=employee, date=date_val).first()
    
    if request.method == 'POST':
        # Form ma'lumotlarini olish
        status = request.POST.get('status')
        comment = request.POST.get('comment', '')
        date_from_form = request.POST.get('date')
        
        # Agar sana form orqali kelsa, uni ishlatish
        if date_from_form:
            try:
                date_val = date.fromisoformat(date_from_form)
            except ValueError:
                date_val = date.today()
        
        # Yopiq kun yoki yakshanba tekshirish (POST uchun ham)
        is_dayoff = DayOff.objects.filter(date=date_val).exists()
        is_sunday = date_val.weekday() == 6
        if is_dayoff or is_sunday:
            dayoff_reason = DayOff.objects.filter(date=date_val).first()
            reason = dayoff_reason.reason if dayoff_reason else "Yakshanba"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'message': f'{date_val.strftime("%d.%m.%Y")} - {reason} kunida davomat o\'zgartirib bo\'lmaydi!'})
            messages.error(request, f"{date_val.strftime('%d.%m.%Y')} - {reason} kunida davomat o'zgartirib bo'lmaydi!")
            return redirect('dashboard')
        
        # Validatsiya
        if not status:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'message': 'Davomat holatini tanlang!'})
            messages.error(request, "Davomat holatini tanlang!")
            form = AttendanceForm(instance=attendance) if attendance else AttendanceForm(initial={'status': 'present'})
            return render(request, 'attendance/individual_attendance_form.html', {
                'form': form,
                'employee': employee,
                'date_val': date_val,
                'dayoff': DayOff.objects.filter(date=date_val).first()
            })
        
        # Agar status absent, sick yoki vacation bo'lsa, izoh majburiy
        if status in ['absent', 'sick', 'vacation'] and not comment.strip():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'message': 'Izoh/sabab kiritish majburiy!'})
            messages.error(request, "Izoh/sabab kiritish majburiy!")
            form = AttendanceForm(instance=attendance) if attendance else AttendanceForm(initial={'status': status})
            return render(request, 'attendance/individual_attendance_form.html', {
                'form': form,
                'employee': employee,
                'date_val': date_val,
                'dayoff': DayOff.objects.filter(date=date_val).first()
            })
        
        # Mavjud davomatni qidirish (date_val yangilanganidan keyin)
        # update_or_create ishlatish - mavjud bo'lsa yangilaydi, yo'q bo'lsa yaratadi
        try:
            attendance, created = Attendance.objects.update_or_create(
                    employee=employee,
                    date=date_val,
                defaults={
                    'status': status,
                    'comment': comment
                }
                )
            
            # AJAX request uchun JSON response qaytarish
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({
                    'success': True,
                    'message': f'{employee} uchun davomat saqlandi!'
                })
            
            messages.success(request, f"{employee} uchun davomat saqlandi!")
            
            # Keyingi xodimga o'tish yoki ro'yxatga qaytish
            next_action = request.POST.get('next_action')
            if next_action == 'next_employee':
                # Xodimlar ro'yxatidan keyingi xodimni topish
                next_employee = Employee.objects.filter(
                    is_active=True,
                    location=employee.location
                ).filter(
                    Q(last_name__gt=employee.last_name) | 
                    (Q(last_name=employee.last_name) & Q(first_name__gt=employee.first_name))
                ).order_by('last_name', 'first_name').first()
                
                if next_employee:
                    return redirect('individual_attendance_create', employee_id=next_employee.id)
                else:
                    messages.info(request, "Siz ushbu joylashuvdagi oxirgi xodimga davomat kiritdingiz.")
                    return redirect('attendance_list')
            
            return redirect('attendance_list')
            
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({'success': False, 'message': f'Xatolik yuz berdi: {str(e)}'})
            messages.error(request, f"Xatolik yuz berdi: {str(e)}")
    
    # GET so'rovi uchun forma tayyorlash
    form = AttendanceForm(instance=attendance) if attendance else AttendanceForm(initial={'status': 'present'})
    
    # Ish kuni emas kunini tekshirish
    dayoff = DayOff.objects.filter(date=date_val).first()
    
    return render(request, 'attendance/individual_attendance_form.html', {
        'form': form,
        'employee': employee,
        'date_val': date_val,
        'dayoff': dayoff
    })

@login_required
def attendance_statistics(request):
    period = request.GET.get('period', 'month')
    employee_id = request.GET.get('employee')
    today = timezone.now().date()
    date_from, date_to = None, None

    if period == 'day':
        date_from = today
        date_to = today
    elif period == 'week':
        date_from = today - timedelta(days=today.weekday())
        date_to = date_from + timedelta(days=6)
    elif period == 'month':
        date_from = today.replace(day=1)
        next_month = (date_from.replace(day=28) + timedelta(days=4)).replace(day=1)
        date_to = next_month - timedelta(days=1)
    elif period == 'quarter':
        month = (today.month - 1) // 3 * 3 + 1
        date_from = date(today.year, month, 1)
        if month == 10:
            date_to = date(today.year, 12, 31)
        else:
            date_to = date(today.year, month + 3, 1) - timedelta(days=1)
    elif period == 'halfyear':
        if today.month <= 6:
            date_from = date(today.year, 1, 1)
            date_to = date(today.year, 6, 30)
        else:
            date_from = date(today.year, 7, 1)
            date_to = date(today.year, 12, 31)
    elif period == 'year':
        date_from = date(today.year, 1, 1)
        date_to = date(today.year, 12, 31)
    elif period == 'custom':
        try:
            date_from = date.fromisoformat(request.GET.get('date_from'))
            date_to = date.fromisoformat(request.GET.get('date_to'))
        except Exception:
            messages.error(request, "Sana formatida xatolik!")
            date_from = today.replace(day=1)
            date_to = today
    else:
        date_from = today.replace(day=1)
        date_to = today

    # Xodim filtri qo'shish
    attendances = Attendance.objects.filter(date__range=[date_from, date_to])
    if employee_id:
        attendances = attendances.filter(employee_id=employee_id)

    stats_by_status = attendances.values('status').annotate(count=Count('id')).order_by('status')
    
    # Calculate total for percentage calculations
    total_attendance = attendances.count()
    if total_attendance > 0:
        for item in stats_by_status:
            item['percentage'] = (item['count'] / total_attendance) * 100
    else:
        for item in stats_by_status:
            item['percentage'] = 0
    stats_by_employee = attendances.values(
        'employee__last_name', 'employee__first_name'
    ).annotate(
        total=Count('id'),
        present=Count('id', filter=Q(status='present')),
        absent=Count('id', filter=Q(status='absent')),
        late=Count('id', filter=Q(status='late')),
        vacation=Count('id', filter=Q(status='vacation')),
        sick=Count('id', filter=Q(status='sick')),
        business=Count('id', filter=Q(status='business')),
    ).order_by('-present', 'employee__last_name')

    stats_by_department = attendances.values(
        'employee__department'
    ).annotate(
        total=Count('id'),
        present=Count('id', filter=Q(status='present')),
        absent=Count('id', filter=Q(status='absent')),
        late=Count('id', filter=Q(status='late')),
        vacation=Count('id', filter=Q(status='vacation')),
        sick=Count('id', filter=Q(status='sick')),
        business=Count('id', filter=Q(status='business')),
    ).order_by('-present')
    
    # Joylashuv bo'yicha statistika
    stats_by_location = attendances.values(
        'employee__location'
    ).annotate(
        total=Count('id'),
        present=Count('id', filter=Q(status='present')),
        absent=Count('id', filter=Q(status='absent')),
        late=Count('id', filter=Q(status='late')),
        vacation=Count('id', filter=Q(status='vacation')),
        sick=Count('id', filter=Q(status='sick')),
        business=Count('id', filter=Q(status='business')),
    ).order_by('-present')

    trend = attendances.values('date', 'status').annotate(count=Count('id')).order_by('date')

    # QuerySet obyektlarini list formatiga o'tkazish
    stats_by_status_list = list(stats_by_status)
    stats_by_employee_list = list(stats_by_employee)
    stats_by_department_list = list(stats_by_department)
    stats_by_location_list = list(stats_by_location)
    trend_list = list(trend)
    
    # Kunlar bo'yicha statistika
    dates = attendances.values('date').annotate(count=Count('id')).order_by('date')
    stats_by_date = []
    for date_item in dates:
        date_stats = {
            'date': date_item['date'],
            'statuses': {}
        }
        for status in ['present', 'absent', 'late', 'vacation', 'sick', 'business']:
            count = attendances.filter(date=date_item['date'], status=status).count()
            date_stats['statuses'][status] = count
        stats_by_date.append(date_stats)
    
    # Prepare JSON-serialized data for charts
    import json
    
    # Status labels and counts for pie chart
    status_labels = [item['status'].title() for item in stats_by_status_list]
    status_counts = [item['count'] for item in stats_by_status_list]
    
    # Department data for bar chart
    department_labels = [item['employee__department'] for item in stats_by_department_list if item['employee__department']]
    department_present = [item['present'] for item in stats_by_department_list if item['employee__department']]
    department_absent = [item['absent'] for item in stats_by_department_list if item['employee__department']]
    department_late = [item['late'] for item in stats_by_department_list if item['employee__department']]
    
    # Location data for bar chart
    location_labels = [dict(Employee.LOCATION_CHOICES).get(item['employee__location'], item['employee__location']) 
                      for item in stats_by_location_list if item['employee__location']]
    location_present = [item['present'] for item in stats_by_location_list if item['employee__location']]
    location_absent = [item['absent'] for item in stats_by_location_list if item['employee__location']]
    location_late = [item['late'] for item in stats_by_location_list if item['employee__location']]
    
    # Trend data for line chart
    trend_dates = sorted(set(item['date'].strftime('%Y-%m-%d') for item in trend_list))
    trend_data = {}
    for status in ['present', 'absent', 'late', 'vacation', 'sick', 'business']:
        trend_data[status] = []
        for date_str in trend_dates:
            date_obj = date.fromisoformat(date_str)
            count = next((item['count'] for item in trend_list if item['date'] == date_obj and item['status'] == status), 0)
            trend_data[status].append(count)
    
    # Xodimlar ro'yxati
    employees = Employee.objects.filter(is_active=True).order_by('last_name', 'first_name')
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'stats_by_status': stats_by_status_list,
        'stats_by_employee': stats_by_employee_list,
        'stats_by_department': stats_by_department_list,
        'stats_by_location': stats_by_location_list,
        'trend': trend_list,
        'stats_by_date': stats_by_date,
        'period': period,
        'employees': employees,
        'selected_employee': employee_id,
        'location_choices': dict(Employee.LOCATION_CHOICES),
        # JSON-serialized data for charts
        'status_labels': json.dumps(status_labels),
        'status_counts': json.dumps(status_counts),
        'department_labels': json.dumps(department_labels),
        'department_present': json.dumps(department_present),
        'department_absent': json.dumps(department_absent),
        'department_late': json.dumps(department_late),
        'location_labels': json.dumps(location_labels),
        'location_present': json.dumps(location_present),
        'location_absent': json.dumps(location_absent),
        'location_late': json.dumps(location_late),
        'trend_dates': json.dumps(trend_dates),
        'trend_data': json.dumps(trend_data),
    }
    return render(request, 'attendance/attendance_statistics.html', context)

@login_required
def salary_statistics_view(request):
    import datetime
    today = datetime.date.today()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))
    # Statistikani hisoblash (agar kerak bo'lsa)
    calculate_monthly_stats(year, month)
    stats = MonthlyEmployeeStat.objects.filter(year=year, month=month).select_related('employee')
    form = SalaryStatFilterForm(initial={'year': year, 'month': month})
    
    # Ишчи кунларни ҳисоблаш
    working_days_in_month, total_days_in_month = calculate_working_days_in_month(year, month)
    
    # Ҳар бир stat объектига ишчи кунлар сонини қўшиш
    for stat in stats:
        stat.working_days_in_month = working_days_in_month
    # Kelmagan kunlar soni va sanalari
    absent_days = {}
    for stat in stats:
        absents = Attendance.objects.filter(
            employee=stat.employee,
            date__year=year,
            date__month=month,
            status='absent'
        ).values_list('date', flat=True)
        stat.absent_count = len(absents)
        stat.absent_dates = list(absents)
    # Umumiy summalar
    total_salary = sum([s.salary for s in stats])
    total_bonus = sum([s.bonus for s in stats])
    total_penalty = sum([s.penalty for s in stats])
    total_accrued = sum([s.accrued for s in stats])
    total_paid = sum([s.paid for s in stats])
    total_debt_start = sum([s.debt_start for s in stats])
    total_debt_end = sum([s.debt_end for s in stats])
    total_absent = sum([s.absent_count for s in stats])
    currency_set = set([s.currency for s in stats])
    total_currency = currency_set.pop() if len(currency_set) == 1 else '...'
    # Valyuta bo'yicha jami qiymatlar
    currency_totals = {}
    for stat in stats:
        cur = stat.currency
        if cur not in currency_totals:
            currency_totals[cur] = {
                'salary': 0, 'bonus': 0, 'penalty': 0, 'accrued': 0, 'paid': 0, 'debt_start': 0, 'debt_end': 0
            }
        currency_totals[cur]['salary'] += float(stat.salary)
        currency_totals[cur]['bonus'] += float(stat.bonus)
        currency_totals[cur]['penalty'] += float(stat.penalty)
        currency_totals[cur]['accrued'] += float(stat.accrued)
        currency_totals[cur]['paid'] += float(stat.paid)
        currency_totals[cur]['debt_start'] += float(stat.debt_start)
        currency_totals[cur]['debt_end'] += float(stat.debt_end)
    return render(request, 'attendance/salary_statistics.html', {
        'stats': stats,
        'form': form,
        'year': year,
        'month': month,
        'total_salary': total_salary,
        'total_bonus': total_bonus,
        'total_penalty': total_penalty,
        'total_accrued': total_accrued,
        'total_paid': total_paid,
        'total_debt_start': total_debt_start,
        'total_debt_end': total_debt_end,
        'absent_days': absent_days,
        'total_absent': total_absent,
        'total_currency': total_currency,
        'currency_totals': currency_totals,
    })

@login_required
def export_salary_statistics_excel(request):
    import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule
    from openpyxl.utils import get_column_letter
    
    today = datetime.date.today()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))
    calculate_monthly_stats(year, month)
    stats = MonthlyEmployeeStat.objects.filter(year=year, month=month).select_related('employee')
    
    # Excel faylini yaratish
    wb = Workbook()
    
    # Ishchi kunlarni hisoblash
    working_days_in_month, total_days_in_month = calculate_working_days_in_month(year, month)
    
    # Ishchi kunlar har bir stat obyektiga biriktiriladi
    for stat in stats:
        stat.working_days_in_month = working_days_in_month

    # Xodimlarni turlariga qarab guruhlash
    half_stats = [s for s in stats if s.employee.employee_type == 'half']
    full_stats = [s for s in stats if s.employee.employee_type == 'full']
    office_stats = [s for s in stats if s.employee.employee_type == 'office']
    weekly_stats = [s for s in stats if s.employee.employee_type == 'weekly']
    guard_stats = [s for s in stats if s.employee.employee_type == 'guard']
    non_office_stats = [s for s in stats if s.employee.employee_type != 'office']
    
    # Chegara va ramkalar stili
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='medium')
    )
    
    # Sodda va toza worksheet yaratish funksiyasi
    def create_worksheet(worksheet, title, data, color_bg='4A90E2', color_accent='2E5C8A'):
        # Sodda sarlavha
        worksheet.merge_cells('A1:H2')
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = f"ISOMER OIL - {title} ({year}-{month:02d})"
        title_cell.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color=color_bg, end_color=color_bg, fill_type='solid')
        
        # Sarlavha uchun oddiy ramka
        for col in range(1, 9):
            for row in range(1, 3):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=color_bg, end_color=color_bg, fill_type='solid')
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
        
        # Sodda sarlavhalar
        headers = [
            "Xodim", "Oylik", "Valyuta", "Kelgan/Jami", 
            "Foiz", "Hisoblangan", "To'langan", "Bonus"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(name='Arial', bold=True, size=11, color='000000')
            cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
        
        # Sodda ma'lumotlar
        for row, stat in enumerate(data, 5):
            # Kelish foizini hisoblash
            if hasattr(stat, 'working_days_in_month') and stat.working_days_in_month > 0:
                percentage = (stat.worked_days / stat.working_days_in_month) * 100
            else:
                percentage = 0
                
            # Ma'lumotlarni formatlash
            row_data = [
                f"{stat.employee.last_name} {stat.employee.first_name}",
                float(stat.salary),
                stat.currency.upper() if stat.currency else 'UZS',
                f"{stat.worked_days}/{stat.working_days_in_month}",
                f"{percentage:.1f}%",
                float(stat.accrued),
                float(stat.paid),
                float(stat.bonus) if stat.bonus else 0
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row, column=col)
                cell.value = value
                
                # Oddiy ramka
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                
                # Sodda stillar
                if col == 1:  # Ism
                    cell.font = Font(name='Arial', size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                elif col == 3:  # Valyuta
                    cell.font = Font(name='Arial', bold=True, size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col == 4:  # Kelgan/Jami
                    cell.font = Font(name='Arial', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col == 5:  # Foiz
                    cell.font = Font(name='Arial', bold=True, size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Foiz uchun oddiy rang berish
                    if percentage >= 80:
                        cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                    elif percentage >= 60:
                        cell.fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid')
                elif col in [2, 6, 7, 8]:  # Pul ustunlari
                    cell.font = Font(name='Arial', size=10)
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Oddiy raqam formati
                    currency = stat.currency.upper() if stat.currency else 'UZS'
                    if currency == 'USD':
                        cell.number_format = '"$" #,##0'
                    elif currency in ['UZS', 'SUM']:
                        cell.number_format = '#,##0 "so\'m"'
                    else:
                        cell.number_format = f'#,##0 "{currency}"'
        
        # Oddiy ustun kengliklari
        column_widths = [25, 15, 10, 15, 12, 18, 15, 12]
        for col, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = width
        
        # Oddiy zebra chiziqlar
        data_end_row = worksheet.max_row
        for row in range(5, data_end_row + 1):
            if row % 2 == 0:  # Juft qatorlar
                for col in range(1, 9):
                    cell = worksheet.cell(row=row, column=col)
                    if not cell.fill or cell.fill.start_color.rgb == '00000000':  # Agar rang berilmagan bo'lsa
                        cell.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
        
        # Sodda valyuta umumiy hisob-kitobi
        if data:
            # Valyutalar bo'yicha guruhlash
            currency_totals = {}
            for stat in data:
                currency = stat.currency.upper() if stat.currency else 'UZS'
                if currency not in currency_totals:
                    currency_totals[currency] = {'count': 0, 'salary': 0, 'accrued': 0, 'paid': 0, 'bonus': 0}
                
                currency_totals[currency]['count'] += 1
                currency_totals[currency]['salary'] += float(stat.salary)
                currency_totals[currency]['accrued'] += float(stat.accrued)
                currency_totals[currency]['paid'] += float(stat.paid)
                currency_totals[currency]['bonus'] += float(stat.bonus) if stat.bonus else 0
            
            # Umumiy ma'lumotlar sarlavhasi
            summary_start_row = data_end_row + 2
            worksheet.merge_cells(f'A{summary_start_row}:H{summary_start_row}')
            summary_header = worksheet.cell(row=summary_start_row, column=1)
            summary_header.value = "VALYUTA BO'YICHA JAMI"
            summary_header.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
            summary_header.fill = PatternFill(start_color=color_bg, end_color=color_bg, fill_type='solid')
            summary_header.alignment = Alignment(horizontal='center', vertical='center')
            summary_header.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Valyuta sarlavhalari
            headers = ["Valyuta", "Soni", "Jami oylik", "Hisoblangan", "To'langan", "Bonus", "Qarzdorlik", ""]
            
            header_row = summary_start_row + 1
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=header_row, column=col)
                cell.value = header
                cell.font = Font(name='Arial', bold=True, size=10, color='000000')
                cell.fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
            
            # Har bir valyuta uchun ma'lumotlar
            current_row = header_row + 1
            for currency, totals in sorted(currency_totals.items()):
                debt_amount = totals['accrued'] - totals['paid']
                
                currency_data = [
                    currency, totals['count'], totals['salary'], totals['accrued'],
                    totals['paid'], totals['bonus'], debt_amount, ""
                ]
                
                for col, value in enumerate(currency_data, 1):
                    cell = worksheet.cell(row=current_row, column=col)
                    cell.value = value
                    cell.font = Font(name='Arial', size=10)
                    cell.alignment = Alignment(
                        horizontal='center' if col in [1, 2] else 'right' if col <= 7 else 'left', 
                        vertical='center'
                    )
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    
                    # Pul ustunlari uchun formatlash
                    if col in [3, 4, 5, 6, 7] and isinstance(value, (int, float)):
                        if currency == 'USD':
                            cell.number_format = '"$" #,##0'
                        elif currency in ['UZS', 'SUM']:
                            cell.number_format = '#,##0 "so\'m"'
                        else:
                            cell.number_format = f'#,##0 "{currency}"'
                
                current_row += 1
        
        # Professional freeze panes va print settings
        worksheet.freeze_panes = 'A5'
    
        # Print settings - professional ko'rinish
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        
        # Print titles (header har sahifada ko'rinadi)
        worksheet.print_title_rows = '1:4'
        
        # Margins
        worksheet.page_margins.left = 0.5
        worksheet.page_margins.right = 0.5
        worksheet.page_margins.top = 0.75
        worksheet.page_margins.bottom = 0.75
        worksheet.page_margins.header = 0.3
        worksheet.page_margins.footer = 0.3
    
    # Guruhlash bilan worksheet yaratish funksiyasi (BARCHA XODIMLAR uchun)
    def create_grouped_worksheet(worksheet, title, all_stats, color_bg='1E3A5F', color_accent='2C5282'):
        # Professional sarlavha - gradient effekt
        worksheet.merge_cells('A1:H2')
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = f"🏢 ISOMER OIL - {title} ({year}-{month:02d})"
        title_cell.font = Font(name='Calibri', bold=True, size=16, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_cell.fill = PatternFill(start_color=color_bg, end_color=color_accent, fill_type='solid')
        
        # Sarlavha uchun professional ramka
        thick_border = Border(
            left=Side(style='thick', color='FFFFFF'),
            right=Side(style='thick', color='FFFFFF'),
            top=Side(style='thick', color='FFFFFF'),
            bottom=Side(style='thick', color='FFFFFF')
        )
        for col in range(1, 9):
            for row in range(1, 3):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=color_bg, end_color=color_accent, fill_type='solid')
                cell.border = thick_border
        
        # Professional ustun sarlavhalari
        headers = [
            "👤 Xodim", "💰 Oylik", "💵 Valyuta", "📅 Kelgan/Jami", 
            "📊 Foiz", "💼 Hisoblangan", "✅ To'langan", "🎁 Bonus"
        ]
        
        header_border = Border(
            left=Side(style='medium', color='1E3A5F'),
            right=Side(style='medium', color='1E3A5F'),
            top=Side(style='medium', color='1E3A5F'),
            bottom=Side(style='thick', color='1E3A5F')
        )
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='5B7FBD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = header_border
        
        # Xodimlarni turlariga qarab guruhlash - professional ranglar
        employee_type_groups = {
            'full': ('⭐ To\'liq stavka xodimlar', '2E75B6'),
            'half': ('📋 15 kunlik xodimlar', 'FF6B35'),
            'weekly': ('📆 Haftada 1 kun xodimlar', '6C757D'),
            'guard': ('🛡️ Qorovul xodimlar', 'DC3545'),
            'office': ('💼 Ofis xodimlari', '8E44AD'),
        }
        
        current_row = 5
        all_currency_totals = {}
        
        # Har bir guruh uchun
        for emp_type, (group_title, group_color) in employee_type_groups.items():
            group_stats = [s for s in all_stats if s.employee.employee_type == emp_type]
            
            if not group_stats:
                continue
            
            # Professional guruh sarlavhasi
            worksheet.merge_cells(f'A{current_row}:H{current_row}')
            group_header = worksheet.cell(row=current_row, column=1)
            group_header.value = group_title.upper()
            group_header.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
            # Gradient effekt uchun biroz qorong'i rang
            darker_color = hex(int(group_color, 16) - 0x202020 if int(group_color, 16) > 0x202020 else 0)[2:].upper().zfill(6)
            group_header.fill = PatternFill(start_color=group_color, end_color=darker_color, fill_type='solid')
            group_header.alignment = Alignment(horizontal='center', vertical='center')
            group_header_border = Border(
                left=Side(style='thick', color='FFFFFF'),
                right=Side(style='thick', color='FFFFFF'),
                top=Side(style='thick', color='FFFFFF'),
                bottom=Side(style='thick', color='FFFFFF')
            )
            group_header.border = group_header_border
            # Row height oshirish
            worksheet.row_dimensions[current_row].height = 25
            current_row += 1
            
            # Guruh xodimlari
            for stat in group_stats:
                # Kelish foizini hisoblash
                if hasattr(stat, 'working_days_in_month') and stat.working_days_in_month > 0:
                    percentage = (stat.worked_days / stat.working_days_in_month) * 100
                else:
                    percentage = 0
                
                # Ma'lumotlarni formatlash
                row_data = [
                    f"{stat.employee.last_name} {stat.employee.first_name}",
                    float(stat.salary),
                    stat.currency.upper() if stat.currency else 'UZS',
                    f"{stat.worked_days}/{stat.working_days_in_month}",
                    f"{percentage:.1f}%",
                    float(stat.accrued),
                    float(stat.paid),
                    float(stat.bonus) if stat.bonus else 0
                ]
                
                # Professional cell border
                data_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                
                for col, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=current_row, column=col)
                    cell.value = value
                    cell.border = data_border
                    
                    # Professional stillar
                    if col == 1:  # Ism
                        cell.font = Font(name='Calibri', size=11, color='1F1F1F')
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif col == 3:  # Valyuta
                        cell.font = Font(name='Calibri', bold=True, size=11, color='1F1F1F')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        # Valyuta badge effekt
                        cell.fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
                    elif col == 4:  # Kelgan/Jami
                        cell.font = Font(name='Calibri', size=11, color='1F1F1F')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col == 5:  # Foiz
                        cell.font = Font(name='Calibri', bold=True, size=11)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        # Professional foiz ranglari
                        if percentage >= 90:
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                            cell.font = Font(name='Calibri', bold=True, size=11, color='006100')
                        elif percentage >= 80:
                            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                            cell.font = Font(name='Calibri', bold=True, size=11, color='9C6500')
                        elif percentage >= 60:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                            cell.font = Font(name='Calibri', bold=True, size=11, color='9C0006')
                        else:
                            cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
                            cell.font = Font(name='Calibri', bold=True, size=11, color='9C0006')
                    elif col in [2, 6, 7, 8]:  # Pul ustunlari
                        cell.font = Font(name='Calibri', size=11, color='1F1F1F')
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        
                        # Professional raqam formati
                        currency = stat.currency.upper() if stat.currency else 'UZS'
                        if currency == 'USD':
                            cell.number_format = '"$" #,##0.00'
                        elif currency in ['UZS', 'SUM']:
                            cell.number_format = '#,##0 "so\'m"'
                        else:
                            cell.number_format = f'#,##0.00 "{currency}"'
                
                # Umumiy valyuta hisob-kitobiga qo'shish
                currency = stat.currency.upper() if stat.currency else 'UZS'
                if currency not in all_currency_totals:
                    all_currency_totals[currency] = {'count': 0, 'salary': 0, 'accrued': 0, 'paid': 0, 'bonus': 0}
                all_currency_totals[currency]['count'] += 1
                all_currency_totals[currency]['salary'] += float(stat.salary)
                all_currency_totals[currency]['accrued'] += float(stat.accrued)
                all_currency_totals[currency]['paid'] += float(stat.paid)
                all_currency_totals[currency]['bonus'] += float(stat.bonus) if stat.bonus else 0
                
                # Professional zebra chiziqlar
                if current_row % 2 == 0:
                    for col in range(1, 9):
                        cell = worksheet.cell(row=current_row, column=col)
                        # Faqat rang berilmagan celllarga rang berish
                        if not hasattr(cell.fill, 'start_color') or cell.fill.start_color.rgb == '00000000' or cell.fill.start_color.rgb == 'FFFFFFFF':
                            cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                
                current_row += 1
                
                # Agar valyuta USD bo'lsa, qo'shimcha qator qo'shish (so'mda)
                currency = stat.currency.upper() if stat.currency else 'UZS'
                if currency == 'USD':
                    USD_TO_UZS_RATE = 12000  # $ kursi
                    salary_in_uzs = float(stat.salary) * USD_TO_UZS_RATE
                    
                    # Qo'shimcha qator ma'lumotlari
                    uzs_row_data = [
                        f"  (so'mda)",  # Xodim nomi o'rniga
                        salary_in_uzs,  # Oylik so'mda
                        'UZS',  # Valyuta
                        "",  # Kelgan/Jami
                        "",  # Foiz
                        "",  # Hisoblangan
                        "",  # To'langan
                        ""   # Bonus
                    ]
                    
                    for col, value in enumerate(uzs_row_data, 1):
                        cell = worksheet.cell(row=current_row, column=col)
                        cell.value = value
                        
                        # Ramka
                        cell.border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thin', color='000000')
                        )
                        
                        # Professional stillar (so'mda qator)
                        if col == 1:  # Ism
                            cell.font = Font(name='Calibri', size=10, italic=True, color='666666')
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        elif col == 2:  # Oylik so'mda
                            cell.font = Font(name='Calibri', size=11, italic=True, color='1F1F1F')
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            cell.number_format = '#,##0 "so\'m"'
                            # Professional rang
                            cell.fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
                        elif col == 3:  # Valyuta
                            cell.font = Font(name='Calibri', bold=True, size=10, italic=True, color='4472C4')
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
                        else:  # Qolgan ustunlar
                            cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                            cell.font = Font(name='Calibri', size=9, color='CCCCCC')
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    current_row += 1
            
            # Guruhlar orasida bo'sh qator
            current_row += 1
        
        # Ustun kengliklari
        column_widths = [25, 15, 10, 15, 12, 18, 15, 12]
        for col, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = width
        
        # Professional umumiy valyuta hisob-kitobi (oxirida)
        if all_currency_totals:
            summary_start_row = current_row + 2  # Bo'sh qator qo'shish
            worksheet.merge_cells(f'A{summary_start_row}:H{summary_start_row}')
            summary_header = worksheet.cell(row=summary_start_row, column=1)
            summary_header.value = "💰 VALYUTA BO'YICHA JAMI (BARCHA XODIMLAR)"
            summary_header.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
            summary_header.fill = PatternFill(start_color='1E3A5F', end_color='2C5282', fill_type='solid')
            summary_header.alignment = Alignment(horizontal='center', vertical='center')
            summary_header.border = Border(
                left=Side(style='thick', color='FFFFFF'),
                right=Side(style='thick', color='FFFFFF'),
                top=Side(style='thick', color='FFFFFF'),
                bottom=Side(style='thick', color='FFFFFF')
            )
            worksheet.row_dimensions[summary_start_row].height = 30
            
            # Professional valyuta sarlavhalari
            headers = ["💵 Valyuta", "🔢 Soni", "💰 Jami oylik", "💼 Hisoblangan", "✅ To'langan", "🎁 Bonus", "📊 Qarzdorlik", ""]
            
            header_row = summary_start_row + 1
            summary_header_border = Border(
                left=Side(style='medium', color='1E3A5F'),
                right=Side(style='medium', color='1E3A5F'),
                top=Side(style='medium', color='1E3A5F'),
                bottom=Side(style='thick', color='1E3A5F')
            )
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=header_row, column=col)
                cell.value = header
                cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
                cell.fill = PatternFill(start_color='5B7FBD', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = summary_header_border
            
            # Har bir valyuta uchun ma'lumotlar
            current_summary_row = header_row + 1
            for currency, totals in sorted(all_currency_totals.items()):
                debt_amount = totals['accrued'] - totals['paid']
                
                currency_data = [
                    currency, totals['count'], totals['salary'], totals['accrued'],
                    totals['paid'], totals['bonus'], debt_amount, ""
                ]
                
                summary_data_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='medium', color='1E3A5F')
                )
                
                for col, value in enumerate(currency_data, 1):
                    cell = worksheet.cell(row=current_summary_row, column=col)
                    cell.value = value
                    cell.border = summary_data_border
                    
                    # Professional stillar
                    if col in [1, 2]:  # Valyuta va Soni
                        cell.font = Font(name='Calibri', bold=True, size=11, color='1F1F1F')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
                    elif col <= 7:  # Pul ustunlari
                        cell.font = Font(name='Calibri', bold=True, size=11, color='1F1F1F')
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        # Professional rang
                        if col == 7:  # Qarzdorlik
                            if debt_amount > 0:
                                cell.fill = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid')
                                cell.font = Font(name='Calibri', bold=True, size=11, color='9C0006')
                            else:
                                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                                cell.font = Font(name='Calibri', bold=True, size=11, color='006100')
                        else:
                            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                        
                        # Professional raqam formati
                        if col in [3, 4, 5, 6, 7] and isinstance(value, (int, float)):
                            if currency == 'USD':
                                cell.number_format = '"$" #,##0.00'
                            elif currency in ['UZS', 'SUM']:
                                cell.number_format = '#,##0 "so\'m"'
                            else:
                                cell.number_format = f'#,##0.00 "{currency}"'
                    else:  # Bo'sh ustun
                        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                current_summary_row += 1
                
                # Agar valyuta USD bo'lsa, qo'shimcha qator qo'shish (so'mda jami)
                if currency == 'USD':
                    USD_TO_UZS_RATE = 12000  # $ kursi
                    salary_total_uzs = totals['salary'] * USD_TO_UZS_RATE
                    accrued_total_uzs = totals['accrued'] * USD_TO_UZS_RATE
                    paid_total_uzs = totals['paid'] * USD_TO_UZS_RATE
                    bonus_total_uzs = totals['bonus'] * USD_TO_UZS_RATE
                    debt_total_uzs = debt_amount * USD_TO_UZS_RATE
                    
                    # Qo'shimcha qator ma'lumotlari (so'mda)
                    uzs_summary_data = [
                        "  (so'mda)",  # Valyuta o'rniga
                        "",  # Soni
                        salary_total_uzs,  # Jami oylik so'mda
                        accrued_total_uzs,  # Hisoblangan so'mda
                        paid_total_uzs,  # To'langan so'mda
                        bonus_total_uzs,  # Bonus so'mda
                        debt_total_uzs,  # Qarzdorlik so'mda
                        ""  # Bo'sh
                    ]
                    
                    summary_data_border_uzs = Border(
                        left=Side(style='thin', color='CCCCCC'),
                        right=Side(style='thin', color='CCCCCC'),
                        top=Side(style='thin', color='CCCCCC'),
                        bottom=Side(style='medium', color='1E3A5F')
                    )
                    
                    for col, value in enumerate(uzs_summary_data, 1):
                        cell = worksheet.cell(row=current_summary_row, column=col)
                        cell.value = value
                        cell.border = summary_data_border_uzs
                        
                        # Professional stillar (so'mda qator)
                        if col == 1:  # Valyuta o'rniga
                            cell.font = Font(name='Calibri', size=10, italic=True, color='666666')
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                            cell.fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
                        elif col == 2:  # Soni
                            cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                        elif col <= 7:  # Pul ustunlari
                            cell.font = Font(name='Calibri', bold=True, size=11, italic=True, color='1F1F1F')
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            cell.fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
                            
                            # Professional raqam formati (so'mda)
                            if col in [3, 4, 5, 6, 7] and isinstance(value, (int, float)):
                                cell.number_format = '#,##0 "so\'m"'
                                
                                # Qarzdorlik uchun rang
                                if col == 7:  # Qarzdorlik
                                    if debt_total_uzs > 0:
                                        cell.fill = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid')
                                        cell.font = Font(name='Calibri', bold=True, size=11, italic=True, color='9C0006')
                                    else:
                                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                                        cell.font = Font(name='Calibri', bold=True, size=11, italic=True, color='006100')
                        else:  # Bo'sh ustun
                            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    
                    current_summary_row += 1
        
        # Professional freeze panes va print settings
        worksheet.freeze_panes = 'A5'
        
        # Print settings - professional ko'rinish
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        
        # Print titles (header har sahifada ko'rinadi)
        worksheet.print_title_rows = '1:4'
        
        # Margins
        worksheet.page_margins.left = 0.5
        worksheet.page_margins.right = 0.5
        worksheet.page_margins.top = 0.75
        worksheet.page_margins.bottom = 0.75
        worksheet.page_margins.header = 0.3
        worksheet.page_margins.footer = 0.3
    
    # Faqat bitta list yaratish - BARCHA XODIMLAR (guruhlash bilan)
    # Default worksheetni o'chirish va yangi yaratish
    wb.remove(wb.active)
    
    if stats:
        ws_all = wb.create_sheet("BARCHA XODIMLAR")
        create_grouped_worksheet(ws_all, "Barcha xodimlar (umumiy)", stats, '2C3E50', '34495E')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"💰_ISOMER_OIL_Oylik_Statistika_{year}_{month:02d}_Professional.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def edit_salary_stat(request, stat_id):
    stat = MonthlyEmployeeStat.objects.get(id=stat_id)
    if request.method == 'POST':
        form = SalaryStatEditForm(request.POST, instance=stat)
        if form.is_valid():
            # Eski oylik va valyutani saqlash
            old_salary = stat.salary
            old_currency = stat.currency
            
            # Save the form first
            form.save()
            
            # Agar oylik yoki valyuta o'zgargan bo'lsa, keyingi oylarga o'tkazish
            if stat.salary != old_salary or stat.currency != old_currency:
                from .services import update_future_months_salary
                update_future_months_salary(stat.employee, stat.salary, stat.currency, stat.year, stat.month)
            
            # Recalculate accrued amount with new salary/bonus values
            from .services import calculate_monthly_stats
            calculate_monthly_stats(stat.year, stat.month)
            
            # AJAX request uchun JSON response qaytarish
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.http import JsonResponse
                return JsonResponse({
                    'success': True,
                    'message': 'Ma\'lumotlar muvaffaqiyatli saqlandi va keyingi oylarga o\'tkazildi!',
                    'redirect_url': request.GET.get('next', '/statistics/salary/')
                })
            
            return redirect(f"{request.GET.get('next', '/statistics/salary/')}")
    else:
        form = SalaryStatEditForm(instance=stat)
    return render(request, 'attendance/edit_salary_stat.html', {'form': form, 'stat': stat})

@login_required
def individual_employee_statistics(request, employee_id):
    """Har bir xodimning individual davomat statistikasi"""
    employee = get_object_or_404(Employee, pk=employee_id)
    
    # Filtrlash parametrlari
    year = request.GET.get('year', date.today().year)
    month = request.GET.get('month', date.today().month)
    
    try:
        year = int(year)
        month = int(month)
    except (ValueError, TypeError):
        year = date.today().year
        month = date.today().month
    
    # Oyning boshidan oxirigacha
    from calendar import monthrange
    start_date = date(year, month, 1)
    end_date = date(year, month, monthrange(year, month)[1])
    
    # Xodimning davomat ma'lumotlari
    attendances = Attendance.objects.filter(
        employee=employee,
        date__range=[start_date, end_date]
    ).order_by('date')
    
    # Yopiq kunlar
    dayoffs = DayOff.objects.filter(
        date__range=[start_date, end_date]
    ).values_list('date', flat=True)
    
    # Kunlik ma'lumotlar yaratish
    daily_data = []
    current_date = start_date
    
    while current_date <= end_date:
        # Yakshanba tekshirish
        is_sunday = current_date.weekday() == 6
        is_dayoff = current_date in dayoffs
        
        # Bu kun uchun davomat ma'lumoti
        attendance = attendances.filter(date=current_date).first()
        
        if attendance:
            status = attendance.status
            comment = attendance.comment or ""
        elif is_sunday:
            status = "sunday"
            comment = "Yakshanba"
        elif is_dayoff:
            status = "dayoff"
            comment = "Yopiq kun"
        else:
            status = "unknown"
            comment = "Ma'lumot yo'q"
        
        daily_data.append({
            'date': current_date,
            'status': status,
            'comment': comment,
            'is_sunday': is_sunday,
            'is_dayoff': is_dayoff,
        })
        
        current_date += timedelta(days=1)
    
    # Statistikalar
    total_days = len(daily_data)
    present = len([d for d in daily_data if d['status'] == 'present'])
    absent = len([d for d in daily_data if d['status'] == 'absent'])
    late = len([d for d in daily_data if d['status'] == 'late'])
    sick = len([d for d in daily_data if d['status'] == 'sick'])
    vacation = len([d for d in daily_data if d['status'] == 'vacation'])
    business = len([d for d in daily_data if d['status'] == 'business'])
    sunday = len([d for d in daily_data if d['status'] == 'sunday'])
    dayoff = len([d for d in daily_data if d['status'] == 'dayoff'])
    unknown = len([d for d in daily_data if d['status'] == 'unknown'])
    
    # Foiz hisoblash
    working_days = total_days - sunday - dayoff
    present_percentage = (present / working_days * 100) if working_days > 0 else 0
    absent_percentage = (absent / working_days * 100) if working_days > 0 else 0
    late_percentage = (late / working_days * 100) if working_days > 0 else 0
    sick_percentage = (sick / working_days * 100) if working_days > 0 else 0
    vacation_percentage = (vacation / working_days * 100) if working_days > 0 else 0
    business_percentage = (business / working_days * 100) if working_days > 0 else 0
    sunday_percentage = (sunday / total_days * 100) if total_days > 0 else 0
    unknown_percentage = (unknown / working_days * 100) if working_days > 0 else 0
    
    # Oy navigatsiyasi
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    # Oy nomlari
    months = [
        (1, 'Yanvar'), (2, 'Fevral'), (3, 'Mart'), (4, 'Aprel'),
        (5, 'May'), (6, 'Iyun'), (7, 'Iyul'), (8, 'Avgust'),
        (9, 'Sentabr'), (10, 'Oktabr'), (11, 'Noyabr'), (12, 'Dekabr')
    ]
    
    context = {
        'employee': employee,
        'daily_data': daily_data,
        'year': year,
        'month': month,
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year': next_year,
        'next_month': next_month,
        'months': months,
        'stats': {
            'present': present,
            'absent': absent,
            'late': late,
            'sick': sick,
            'vacation': vacation,
            'business': business,
            'sunday': sunday,
            'dayoff': dayoff,
            'unknown': unknown,
            'total_days': total_days,
            'working_days': working_days,
            'present_percentage': round(present_percentage, 1),
            'absent_percentage': round(absent_percentage, 1),
            'late_percentage': round(late_percentage, 1),
            'sick_percentage': round(sick_percentage, 1),
            'vacation_percentage': round(vacation_percentage, 1),
            'business_percentage': round(business_percentage, 1),
            'sunday_percentage': round(sunday_percentage, 1),
            'unknown_percentage': round(unknown_percentage, 1),
        }
    }
    
    return render(request, 'attendance/individual_employee_statistics.html', context)

@login_required
def edit_attendance_history(request):
    """Avvalgi davomatlarni tahrirlash"""
    # Filtrlash parametrlari
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    employee_id = request.GET.get('employee')
    department = request.GET.get('department')
    
    # Default qiymatlar
    if not date_from:
        date_from = (date.today() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = date.today().strftime('%Y-%m-%d')
    
    # Filtrlash
    filters = {}
    if date_from and date_to:
        try:
            date_from_obj = date.fromisoformat(date_from)
            date_to_obj = date.fromisoformat(date_to)
            filters['date__range'] = (date_from_obj, date_to_obj)
        except ValueError:
            messages.error(request, "Sana formati noto'g'ri!")
            date_from = (date.today() - timedelta(days=30)).strftime('%Y-%m-%d')
            date_to = date.today().strftime('%Y-%m-%d')
            filters['date__range'] = (date.today() - timedelta(days=30), date.today())
    
    if employee_id:
        filters['employee_id'] = employee_id
    if department:
        filters['employee__department'] = department
    
    # Davomat ma'lumotlarini olish
    attendances = Attendance.objects.filter(**filters).select_related('employee').order_by('-date', 'employee__last_name')
    
    # Formset uchun tayyorlash
    AttendanceFormSet = modelformset_factory(
        Attendance, 
        form=AttendanceForm, 
        can_delete=True,
        extra=0,
        fields=['employee', 'date', 'status', 'comment']
    )
    
    if request.method == 'POST':
        formset = AttendanceFormSet(request.POST, queryset=attendances)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Davomat ma'lumotlari yangilandi!")
            return redirect('edit_attendance_history')
    else:
        formset = AttendanceFormSet(queryset=attendances)
    
    # Filtrlash uchun ma'lumotlar
    employees = Employee.objects.filter(is_active=True).order_by('last_name', 'first_name')
    departments = Employee.objects.values_list('department', flat=True).distinct().exclude(department__isnull=True).exclude(department='')
    
    context = {
        'formset': formset,
        'employees': employees,
        'departments': departments,
        'date_from': date_from,
        'date_to': date_to,
        'selected_employee': employee_id,
        'selected_department': department,
        'attendances': attendances,
    }
    
    return render(request, 'attendance/edit_attendance_history.html', context)

def employee_attendance_history(request, employee_id):
    """Xodimning kunlik davomat tarixini ko'rsatish"""
    employee = get_object_or_404(Employee, id=employee_id, is_active=True)
    
    # Filtrlash parametrlari
    year = request.GET.get('year', date.today().year)
    month = request.GET.get('month', date.today().month)
    
    try:
        year = int(year)
        month = int(month)
    except (ValueError, TypeError):
        year = date.today().year
        month = date.today().month
    
    # Oyning boshidan oxirigacha
    from calendar import monthrange
    start_date = date(year, month, 1)
    end_date = date(year, month, monthrange(year, month)[1])
    
    # Xodimning bu oydagi davomat ma'lumotlari
    attendance_records = Attendance.objects.filter(
        employee=employee,
        date__range=[start_date, end_date]
    ).order_by('date')
    
    # Yopiq kunlar
    dayoffs = DayOff.objects.filter(
        date__range=[start_date, end_date]
    ).values_list('date', flat=True)
    
    # Kalendar kunlari yaratish (7x7 grid uchun)
    calendar_days = []
    current_date = start_date
    
    # Oyning birinchi kunidan oldingi bo'sh kunlar
    first_weekday = start_date.weekday()  # 0 = Dushanba, 6 = Yakshanba
    for i in range(first_weekday):
        calendar_days.append({
            'day': '',
            'status': 'empty',
            'status_class': 'empty',
            'status_text': '',
            'comment': '',
            'date': None
        })
    
    # Oy kunlari
    while current_date <= end_date:
        # Yakshanba tekshirish
        is_sunday = current_date.weekday() == 6
        is_dayoff = current_date in dayoffs
        
        # Bu kun uchun davomat ma'lumoti
        attendance = attendance_records.filter(date=current_date).first()
        
        if attendance:
            status = attendance.status
            comment = attendance.comment or ""
        elif is_sunday:
            status = "sunday"
            comment = "Yakshanba"
        elif is_dayoff:
            status = "dayoff"
            comment = "Yopiq kun"
        else:
            status = "unknown"
            comment = "Ma'lumot yo'q"
        
        # Status class va text
        status_class = status
        status_text = get_status_text(status)
        
        calendar_days.append({
            'day': current_date.day,
            'status': status,
            'status_class': status_class,
            'status_text': status_text,
            'comment': comment,
            'date': current_date
        })
        
        current_date += timedelta(days=1)
    
    # Oyning oxiridan keyingi bo'sh kunlar
    last_weekday = end_date.weekday()
    remaining_days = 6 - last_weekday
    for i in range(remaining_days):
        calendar_days.append({
            'day': '',
            'status': 'empty',
            'status_class': 'empty',
            'status_text': '',
            'comment': '',
            'date': None
        })
    
    # Statistikalar
    total_days = len([d for d in calendar_days if d['date']])
    present = len([d for d in calendar_days if d['status'] == 'present'])
    absent = len([d for d in calendar_days if d['status'] == 'absent'])
    late = len([d for d in calendar_days if d['status'] == 'late'])
    sick = len([d for d in calendar_days if d['status'] == 'sick'])
    vacation = len([d for d in calendar_days if d['status'] == 'vacation'])
    business = len([d for d in calendar_days if d['status'] == 'business'])
    sunday = len([d for d in calendar_days if d['status'] == 'sunday'])
    dayoff = len([d for d in calendar_days if d['status'] == 'dayoff'])
    unknown = len([d for d in calendar_days if d['status'] == 'unknown'])
    
    # Foiz hisoblash
    working_days = total_days - sunday - dayoff
    present_percentage = (present / working_days * 100) if working_days > 0 else 0
    absent_percentage = (absent / working_days * 100) if working_days > 0 else 0
    late_percentage = (late / working_days * 100) if working_days > 0 else 0
    sick_percentage = (sick / working_days * 100) if working_days > 0 else 0
    vacation_percentage = (vacation / working_days * 100) if working_days > 0 else 0
    business_percentage = (business / working_days * 100) if working_days > 0 else 0
    sunday_percentage = (sunday / total_days * 100) if total_days > 0 else 0
    unknown_percentage = (unknown / working_days * 100) if working_days > 0 else 0
    
    # Oy navigatsiyasi
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    # Oy nomlari
    months = [
        (1, 'Yanvar'), (2, 'Fevral'), (3, 'Mart'), (4, 'Aprel'),
        (5, 'May'), (6, 'Iyun'), (7, 'Iyul'), (8, 'Avgust'),
        (9, 'Sentabr'), (10, 'Oktabr'), (11, 'Noyabr'), (12, 'Dekabr')
    ]
    
    context = {
        'employee': employee,
        'calendar_days': calendar_days,
        'year': year,
        'month': month,
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year': next_year,
        'next_month': next_month,
        'months': months,
        'stats': {
            'present': present,
            'absent': absent,
            'late': late,
            'sick': sick,
            'vacation': vacation,
            'business': business,
            'sunday': sunday,
            'dayoff': dayoff,
            'unknown': unknown,
            'total_days': total_days,
            'working_days': working_days,
            'present_percentage': round(present_percentage, 1),
            'absent_percentage': round(absent_percentage, 1),
            'late_percentage': round(late_percentage, 1),
            'sick_percentage': round(sick_percentage, 1),
            'vacation_percentage': round(vacation_percentage, 1),
            'business_percentage': round(business_percentage, 1),
            'sunday_percentage': round(sunday_percentage, 1),
            'unknown_percentage': round(unknown_percentage, 1),
        }
    }
    
    return render(request, 'attendance/employee_attendance_history.html', context)

@login_required
def get_attendance_data_ajax(request, employee_id):
    """AJAX orqali davomat ma'lumotlarini olish"""
    from django.http import JsonResponse
    
    employee = get_object_or_404(Employee, id=employee_id, is_active=True)
    date_str = request.GET.get('date')
    
    if not date_str:
        return JsonResponse({'success': False, 'message': 'Sana ko\'rsatilmagan'})
    
    try:
        date_val = date.fromisoformat(date_str)
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Noto\'g\'ri sana formati'})
    
    # Davomat ma'lumotini olish
    attendance = Attendance.objects.filter(employee=employee, date=date_val).first()
    
    # Yopiq kun yoki yakshanba tekshirish
    is_dayoff = DayOff.objects.filter(date=date_val).exists()
    is_sunday = date_val.weekday() == 6
    
    data = {
        'success': True,
        'employee_id': employee.id,
        'employee_name': f"{employee.last_name} {employee.first_name}",
        'date': date_str,
        'status': attendance.status if attendance else '',
        'comment': attendance.comment if attendance else '',
        'is_dayoff': is_dayoff,
        'is_sunday': is_sunday,
        'dayoff_reason': DayOff.objects.filter(date=date_val).first().reason if is_dayoff else None
    }
    
    return JsonResponse(data)

def get_status_text(status):
    """Status matnini olish"""
    status_map = {
        'present': 'Keldi',
        'absent': 'Kelmadi',
        'late': 'Kechikdi',
        'sick': 'Kasal',
        'vacation': 'Ta\'til',
        'business': 'Ish safari',
        'sunday': 'Yakshanba',
        'dayoff': 'Yopiq kun',
        'unknown': 'Ma\'lumot yo\'q',
        'empty': ''
    }
    return status_map.get(status, 'Noma\'lum')